import json
from openpyxl import Workbook
import matplotlib.pyplot as plt
import numpy as np
from utils.get_intersecting_entries import get_intersecting_entries

greater_arrs = [
    ("bn", "outputs\\hits_lang_comparisons\\en_bn\\hits_en_less_than_bn.json"),
    ("da", "outputs\\hits_lang_comparisons\\en_da\\hits_en_less_than_da.json"),
    ("fi", "outputs\\hits_lang_comparisons\\en_fi\\hits_en_less_than_fi.json")
]

less_arrs = [("bn", "outputs\hits_lang_comparisons\en_bn\hits_en_greater_than_bn.json"),
                ("da", "outputs\hits_lang_comparisons\en_da\hits_en_greater_than_da.json"),
                ("fi", "outputs\hits_lang_comparisons\en_fi\hits_en_greater_than_fi.json")]

same_arrs = [("bn", "outputs\hits_lang_comparisons\en_bn\hits_en_equal_to_bn.json"), ("da", "outputs\hits_lang_comparisons\en_da\hits_en_equal_to_da.json"),
                ("fi", "outputs\hits_lang_comparisons\en_fi\hits_en_equal_to_fi.json")]

c_count = {}
total_count = {"fi":{}, "da":{}, "bn":{}}

def get_comp_count(arrs, t):
    new_data = {}
    with open("data\mintaka_test_extended2.json", "r", encoding="utf-8") as file:
        data = json.load(file)
        intersect = get_intersecting_entries(data)
        id_obj_map = {d["id"]: d for d in data}

        for lang, c_arr in arrs:
            if t not in total_count[lang]:
                total_count[lang][t] = 0
            new_data[lang] = {}
            with open(c_arr, "r", encoding="utf-8") as file:
                comp_data = json.load(file)
                count = 0
                for id, d_entry in comp_data.items():
                    count+=1
                    min_d = id_obj_map[id]
                    complx = min_d["complexityType"]
                    if complx not in new_data[lang]:
                        new_data[lang][complx] = 0
                    new_data[lang][complx] += 1
                    total_count[lang][t] += 1
                print(f"Total {t} for {lang}: count = {count} : file = {c_arr}")
    return new_data


greater_data = get_comp_count(greater_arrs, "greater")
less_data = get_comp_count(less_arrs, "less")
same_ars = get_comp_count(same_arrs, "same")
# Combine greater and less data
new_data = {}
for lang in greater_data.keys():
    new_data[lang] = {}
    for complx_type in greater_data[lang].keys():
        new_data[lang][complx_type] = greater_data[lang].get(complx_type, 0) - less_data[lang].get(complx_type, 0)


# Create Excel workbook
wb = Workbook()
# Remove the default sheet created by openpyxl
wb.remove(wb.active)
headers = ["generic", "multihop", "intersection", "difference", "comparative", "superlative", "ordinal", "count", "yesno"]

def write_data_to_sheet(wb, sheet_name, data):
    ws = wb.create_sheet(title=sheet_name)
    ws.append(["Language"] + headers)
    for lang in data.keys():
        row = [lang, ]
        for h in headers:
            if h in data[lang]:
                row.append(data[lang][h])
            else:
                row.append(0)
        ws.append(row)



write_data_to_sheet(wb, "Greater Data", greater_data)
write_data_to_sheet(wb, "Less Data", less_data)
write_data_to_sheet(wb, "Same Data", same_ars)
write_data_to_sheet(wb, "Difference (Greater-Less)", new_data)
print("Total counts per language:")
for lang, counts in total_count.items():
    print(f"{lang}: {counts}")
# Save workbook
wb.save("complexity_counts2.xlsx")


""" for lang in greater_data.keys():
    new_data[lang] = {}
    for complx_type in greater_data[lang].keys():
        new_data[lang][complx_type] = greater_data[lang].get(complx_type, 0) - less_data[lang].get(complx_type, 0)

def write_data_to_sheet(wb, sheet_name, data):
    ws = wb.create_sheet(title=sheet_name)
    # Collect all possible complexity types for columns
    all_types = set()
    for lang_dict in data.values():
        all_types.update(lang_dict.keys())
    all_types = sorted(all_types)
    # Write header
    header = ["Language"] + all_types
    ws.append(header)
    # Write data rows
    for lang, counts in data.items():
        row = [lang] + [counts.get(t, 0) for t in all_types]
        ws.append(row)

# Create Excel workbook
wb = Workbook()
# Remove the default sheet created by openpyxl
wb.remove(wb.active)

write_data_to_sheet(wb, "Greater Data", greater_data)
write_data_to_sheet(wb, "Less Data", less_data)
write_data_to_sheet(wb, "Same Data", same_ars)
write_data_to_sheet(wb, "Difference (Greater-Less)", new_data)

# Save workbook
wb.save("complexity_counts.xlsx")
"""
# Prepare data for plotting
complexity_types = sorted({ct for lang_dict in new_data.values() for ct in lang_dict.keys()})
languages = list(new_data.keys())

# Gather the values for each language and complexity type
values = []
for lang in languages:
    values.append([new_data[lang].get(ct, 0) for ct in complexity_types])

# Plotting
x = np.arange(len(complexity_types))  # the label locations
width = 0.2  # the width of the bars

fig, ax = plt.subplots(figsize=(12, 6))
bars = []
for i, lang in enumerate(languages):
    bar = ax.bar(x + i*width, values[i], width, label=lang)
    bars.append(bar)

# Add count labels on each bar
for bar_group in bars:
    for rect in bar_group:
        height = rect.get_height()
        ax.annotate(f'{int(height)}',
                    xy=(rect.get_x() + rect.get_width() / 2, height),
                    xytext=(0, 3),  # 3 points vertical offset
                    textcoords="offset points",
                    ha='center', va='bottom', fontsize=9)

# Add labels, title, and legend
ax.set_xlabel('Complexity Type')
ax.set_ylabel('Difference (Better-Worse)')
ax.set_title('Difference by Complexity Type and Language in comparison to English')
ax.set_xticks(x + width)
ax.set_xticklabels(complexity_types, rotation=45)
ax.legend()

plt.tight_layout()
plt.savefig("complexity_difference_barplot.png")
plt.show()


