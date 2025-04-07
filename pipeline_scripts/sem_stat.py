import openpyxl
import json
from openpyxl import Workbook
import matplotlib.pyplot as plt
from collections import Counter
import plotly.graph_objects as go
import os
from utils.get_generation_path import get_generation_path


def calculate_mean(array):
    """Calculate the mean of a list of numbers."""
    if not array:
        return 0
    return sum(array) / len(array)

def calculate_variance(array):
    """Calculate the variance of a list of numbers."""
    if not array:
        return 0
    mean = calculate_mean(array)
    return sum((x - mean) ** 2 for x in array) / len(array)

# Flatten a list of lists
def flatten(array_of_arrays):
    return [item for sublist in array_of_arrays for item in sublist]

def init_id_hit_map(hits_data):
    """Initialize a dictionary mapping IDs to their first hit."""
    id_hit_map = {}
    for id, obj in hits_data.items():
        if id not in id_hit_map:
            id_hit_map[id] = obj["hits"][0]["hit"]
    return id_hit_map

def init_id_variables_map(mintaka_data):
        id_variables_map = {}
        for entry in mintaka_data:
            id = entry["id"]
            complexity = entry["complexityType"]
            category = entry["category"]
            answerType = entry["answer"]["answerType"]
            got_supporting_ents = "answer" in entry and "supportingEnt" in entry["answer"]
            id_variables_map[id] = {
                "complexity": complexity,
                "category": category,
                "answerType": answerType,
                "got_supporting_ents": got_supporting_ents
            }
        return id_variables_map

def run_semantic_similarity_analysis(lang, test_type):
    hits_path = get_generation_path("hit_annotation_json", test_type, lang)
    sem_scores_path = get_generation_path("sem_scores_json", test_type, lang)
    mintaka_path = get_generation_path("test_data_extended", test_type, lang)

    with open(hits_path, "r", encoding="utf-8") as f:
        hits_data = json.load(f)

    with open(sem_scores_path, 'r', encoding="utf-8") as f:
        sem_data = json.load(f)

    with open(mintaka_path, 'r', encoding="utf-8") as f:
        mintaka_data = json.load(f)

    id_hit_map = init_id_hit_map(hits_data)
    
    id_variables_map = init_id_variables_map(mintaka_data)

    def calculate_data():
        sem_scores_obj = {
            "complexity": {},
            "category": {},
            "answerType": {},
            "got_supporting_ents": {}
        }
        worst_performers = []
        best_performers = []

        def check_add_to_worst_performers(entry):
            """Add an entry to the worst performers list."""
            if len(worst_performers) < 10:
                worst_performers.append(entry)
            else:
                worst_performers.sort(key=lambda x: x["mean"])
                if entry["mean"] < worst_performers[-1]["mean"]:
                    worst_performers[-1] = entry

        def check_add_to_best_performers(entry):
            """Add an entry to the best performers list."""
            if len(best_performers) < 10:
                best_performers.append(entry)
            else:
                best_performers.sort(key=lambda x: x["mean"], reverse=True)
                if entry["mean"] > best_performers[-1]["mean"]:
                    best_performers[-1] = entry

        hits_sem_scores = {
            "hit" : [],
            "not_hit": []
        }

        # Process each entry in sem_data 
        for entry in sem_data:
            sem_scores = entry["sem_scores"]
            id = entry["id"]
            variables = id_variables_map[id]
            entry_mean = calculate_mean(sem_scores)
            hit_1 = id_hit_map[id]
            if hit_1 is True:
                hits_sem_scores["hit"].append(entry_mean)
            else:
                hits_sem_scores["not_hit"].append(entry_mean)
            full_entry = {**entry, **variables, "mean": entry_mean, "hit_1": hit_1}
            check_add_to_best_performers(full_entry)
            check_add_to_worst_performers(full_entry)

            for key, value in variables.items():
                if value not in sem_scores_obj[key]:
                    sem_scores_obj[key][value] = []
                sem_scores_obj[key][value].append(sem_scores)
        return sem_scores_obj, best_performers, worst_performers, hits_sem_scores

    def create_semscore_analysis_workbook(sem_scores_obj, best_performers, worst_performers, hits_sem_scores):

        def create_mean_variance_by_key_sheet(wb, sem_scores_obj):
            """Create a sheet with mean and variance by key."""
            # Create a new sheet
            ws = wb.active
            ws.title = "SemScores Analysis"

            # Write headers to the Excel sheet
            ws.append(["Key", "Value", "Mean", "Variance"])

            # Process sem_scores_obj
            for key, value_dict in sem_scores_obj.items():
                for value, sem_scores_list in value_dict.items():

                    # Flatten the sem_scores list
                    flattened_scores = flatten(sem_scores_list)
                    # Compute mean and variance
                    mean = calculate_mean(flattened_scores)
                    variance = calculate_variance(flattened_scores)
                    # Write to Excel
                    ws.append([key, value, mean, variance])

        

        def write_performers_to_sheet(wb, sheet_title, performers):
            """Write performer data to an Excel sheet."""
            wb.create_sheet(title=sheet_title)
            ws = wb[sheet_title]

            # Dynamically write headers based on keys of the first entry in performers
            if performers:
                headers = []
                for key in performers[0].keys():
                    if isinstance(performers[0][key], list):
                        # Expand array values into separate columns
                        headers.extend([f"{key}_{i}" for i in range(len(performers[0][key]))])
                    else:
                        headers.append(key)
                ws.append(headers)

            # Write data rows for performers
            for performer in performers:
                row = []
                for key in performer.keys():
                    if isinstance(performer[key], list):
                        # Expand array values into separate columns
                        row.extend([str(item) for item in performer[key]])
                    else:
                        row.append(str(performer[key]))
                ws.append(row)
        def create_hits_semscore_sheet(wb, sheet_title, hits_sem_scores):
            """Write Hits Sem Scores to an Excel sheet."""
            wb.create_sheet(title=sheet_title)
            ws = wb[sheet_title]
            ws.append(["Hit_1/Not Hit_1", "Mean semscore", "Variance"])
            for key, sem_scores in hits_sem_scores.items():
                mean = calculate_mean(sem_scores)
                variance = calculate_variance(sem_scores)
                ws.append([key, mean, variance])

        """Create an Excel workbook with sem scores analysis."""
        # Create a new workbook
        wb = Workbook()
        # Create a sheet for mean and variance by key
        create_mean_variance_by_key_sheet(wb, sem_scores_obj)
        # Write Best Performers to Excel
        write_performers_to_sheet(wb, "Best Performers", best_performers)
        # Write Worst Performers to Excel
        write_performers_to_sheet(wb, "Worst Performers", worst_performers)
        # Write Hits Sem Scores to Excel
        create_hits_semscore_sheet(wb, "Hits_1 Semscores", hits_sem_scores)
        # Save the workbook
        wb_path = get_generation_path("sem_scores_analysis_sheet", test_type, lang)
        wb.save(wb_path)

    def create_plots(lang = lang, test_type = test_type):
        def group_scores_by_label(data): 
            # Group scores by labels
            grouped_scores = {}
            for score, label in data:
                if label not in grouped_scores:
                    grouped_scores[label] = []
                grouped_scores[label].append(round(score, 2))  # Round scores to 2 decimals
            return grouped_scores   


        def get_sem_scores_of_variable_type(variable_type, sem_data = sem_data, id_variables_map = id_variables_map):
            all_sem_scores = []
            for entry in sem_data:
                sem_scores = entry["sem_scores"]
                id = entry["id"]
                variables = id_variables_map[id]
                entry_mean = calculate_mean(sem_scores)
                variable = variables[variable_type]
                all_sem_scores.append((entry_mean, variable))
            return all_sem_scores

        def ensure_directory_exists(path):
            """Ensure that the directory exists, creating it if necessary."""
            os.makedirs(path, exist_ok=True)

        def create_sem_score_stacked_bar_chart(data, title, variable):
            grouped_scores = group_scores_by_label(data)    
            # Count frequencies for each label

            # Define a set of distinct colors for better contrast
            distinct_colors = [
                "#1f77b4", "#ff7f0e", "#2ca02c", "#d62728", "#9467bd",
                "#ffff00", "#e377c2", "#66ffff"
            ]

            labels = list(grouped_scores.keys())
            # Prepare data for a stacked bar chart
            x_values = sorted(set(score for scores in grouped_scores.values() for score in scores))
            stacked_data = {label: [grouped_scores[label].count(x) for x in x_values] for label in labels}

            # Plot the stacked bar chart
            plt.figure(figsize=(12, 8))
            bottom_values = [0] * len(x_values)  # Initialize the bottom of the stack

            for i, label in enumerate(labels):
                plt.bar(
                    x_values, stacked_data[label], width=0.01,
                    color=distinct_colors[i % len(distinct_colors)],  # Use distinct colors
                    alpha=0.7, label=label, bottom=bottom_values
                )
                # Update the bottom values for the next stack
                bottom_values = [bottom + current for bottom, current in zip(bottom_values, stacked_data[label])]

            plt.xlabel("Sem Score (rounded to 2 decimals)")
            plt.ylabel("Frequency")
            plt.title(f"{title} (Stacked)")
            plt.legend(title="Categories")
            plt.grid(axis='y', linestyle='--', alpha=0.7)
            plt.tight_layout()

            # Ensure the directory exists
            output_dir = f"outputs/sem_score_plots/{test_type}/{lang}/"
            ensure_directory_exists(output_dir)
            plot_path = get_generation_path("sem_scores_stacked_bars_plot", test_type, lang).replace("CATEGORY", variable)
            # Save the plot as an image
            plt.savefig(plot_path)
            print(f"Saved stacked bar chart as {plot_path}")
            # plt.show()

        def create_sem_score_interactive_plot(data, title, variable):
            grouped_scores = group_scores_by_label(data)
            x_values = sorted(set(score for scores in grouped_scores.values() for score in scores))
            traces = []
            labels = list(grouped_scores.keys())

            for label in labels:
                trace = go.Bar(
                    x=x_values,
                    y=[grouped_scores[label].count(x) for x in x_values],
                    name=label
                )
                traces.append(trace)

            # Create the figure with dropdown menu
            fig = go.Figure()

            # Add traces for each category
            for trace in traces:
                fig.add_trace(trace)

            # Update layout with dropdown menu
            fig.update_layout(
                updatemenus=[
                    {
                        "buttons": [
                            {
                                "label": "All Categories",
                                "method": "update",
                                "args": [{"visible": [True] * len(labels)}, {"title": "All Categories"}],
                            }
                        ]
                        + [
                            {
                                "label": label,
                                "method": "update",
                                "args": [
                                    {"visible": [i == j for i in range(len(labels))]},
                                    {"title": f"Category: {label}"},
                                ],
                            }
                            for j, label in enumerate(labels)
                        ],
                        "direction": "down",
                        "showactive": True,
                    }
                ],
                title="Distribution of Sem Scores by Category (Interactive)",
                xaxis_title="Sem Score (rounded to 2 decimals)",
                yaxis_title="Frequency",
            )

            # Ensure the directory exists
            output_dir = f"./sem_score_plots/{test_type}/{lang}/"
            ensure_directory_exists(output_dir)

            # Save the interactive plot as an HTML file
            plot_path = get_generation_path("sem_scores_interactive_plot", test_type, lang).replace("CATEGORY", variable)
            fig.write_html(plot_path)
            print(f"Saved interactive plot as {plot_path}")

            # Show the interactive plot
            # fig.show()

        variables_to_graph = ["complexity", "category", "answerType", "got_supporting_ents"]
        for variable in variables_to_graph:
            sem_scores = get_sem_scores_of_variable_type(variable)
            title = f"Distribution of Sem Scores by {variable}"
            # Create the stacked bar chart
            create_sem_score_stacked_bar_chart(sem_scores, title, variable)
            # Create the interactive plot
            create_sem_score_interactive_plot(sem_scores, title, variable)
    # running the different parts of the analysis
    sem_scores_obj, best_performers, worst_performers, hits_sem_scores = calculate_data()
    create_semscore_analysis_workbook(sem_scores_obj, best_performers, worst_performers, hits_sem_scores)
    create_plots() # Creates stacked bar charts and interactive plots for the sem scores









