import openpyxl
import json
from openpyxl import Workbook
import matplotlib.pyplot as plt
from collections import Counter

def calculate_mean(array):
    """Calculate the mean of a list of numbers."""
    if not array:
        return 0
    return sum(array) / len(array)

def calculate_variance(array):
    """Calculate the variance of a list of numbers."""
    if not array:
        return 0
    mean = calculate_mean(array)
    return sum((x - mean) ** 2 for x in array) / len(array)

# Flatten a list of lists
def flatten(array_of_arrays):
    return [item for sublist in array_of_arrays for item in sublist]


with open("hits_da.json", "r", encoding="utf-8") as f:
    hits_data = json.load(f)

def init_id_hit_map(hits_data):
    """Initialize a dictionary mapping IDs to their first hit."""
    id_hit_map = {}
    for id, obj in hits_data.items():
        if id not in id_hit_map:
            id_hit_map[id] = obj["hits"][0]["hit"]
    return id_hit_map

id_hit_map = init_id_hit_map(hits_data)

with open('sem_test2.json', 'r', encoding="utf-8") as f:
    sem_data = json.load(f)

with open('./data/mintaka_test_extended.json', 'r', encoding="utf-8") as f:
    mintaka_data = json.load(f)

id_variables_map = {}
for entry in mintaka_data:
    id = entry["id"]
    complexity = entry["complexityType"]
    category = entry["category"]
    answerType = entry["answer"]["answerType"]
    got_supporting_ents = "answer" in entry and "supportingEnt" in entry["answer"]
    id_variables_map[id] = {
        "complexity": complexity,
        "category": category,
        "answerType": answerType,
        "got_supporting_ents": got_supporting_ents
    }

sem_scores_obj = {
    "complexity": {},
    "category": {},
    "answerType": {},
    "got_supporting_ents": {}
}

worst_performers = []
best_performers = []

def check_add_to_worst_performers(entry):
    """Add an entry to the worst performers list."""
    if len(worst_performers) < 10:
        worst_performers.append(entry)
    else:
        worst_performers.sort(key=lambda x: x["mean"])
        if entry["mean"] < worst_performers[-1]["mean"]:
            worst_performers[-1] = entry

def check_add_to_best_performers(entry):
    """Add an entry to the best performers list."""
    if len(best_performers) < 10:
        best_performers.append(entry)
    else:
        best_performers.sort(key=lambda x: x["mean"], reverse=True)
        if entry["mean"] > best_performers[-1]["mean"]:
            best_performers[-1] = entry

hits_sem_scores = {
    "hit" : [],
    "not_hit": []
}


all_sem_scores = []

for entry in sem_data:
    sem_scores = entry["sem_scores"]

    id = entry["id"]
    variables = id_variables_map[id]
    entry_mean = calculate_mean(sem_scores)
    all_sem_scores.append(entry_mean)
    hit_1 = id_hit_map[id]
    if hit_1 is True:
        hits_sem_scores["hit"].append(entry_mean)
    else:
        hits_sem_scores["not_hit"].append(entry_mean)
    full_entry = {**entry, **variables, "mean": entry_mean, "hit_1": hit_1}
    check_add_to_best_performers(full_entry)
    check_add_to_worst_performers(full_entry)

    for key, value in variables.items():
        if value not in sem_scores_obj[key]:
            sem_scores_obj[key][value] = []
        sem_scores_obj[key][value].append(sem_scores)




# Create a new workbook and sheet
wb = Workbook()
ws = wb.active
ws.title = "SemScores Analysis"

# Write headers to the Excel sheet
ws.append(["Key", "Value", "Mean", "Variance"])

# Distribution graph, 


# Process sem_scores_obj
for key, value_dict in sem_scores_obj.items():
    for value, sem_scores_list in value_dict.items():

        # Flatten the sem_scores list
        flattened_scores = flatten(sem_scores_list)
        # Compute mean and variance
        mean = calculate_mean(flattened_scores)
        variance = calculate_variance(flattened_scores)
        # Write to Excel
        ws.append([key, value, mean, variance])

# Create a new sheet for Best Performers
wb.create_sheet(title="Best Performers")
ws = wb["Best Performers"]

# Dynamically write headers based on keys of the first entry in best_performers
if best_performers:
    headers = []
    for key in best_performers[0].keys():
        if isinstance(best_performers[0][key], list):
            # Expand array values into separate columns
            headers.extend([f"{key}_{i}" for i in range(len(best_performers[0][key]))])
        else:
            headers.append(key)
    ws.append(headers)

# Write data rows for Best Performers
for performer in best_performers:
    row = []
    for key in performer.keys():
        if isinstance(performer[key], list):
            # Expand array values into separate columns
            row.extend([str(item) for item in performer[key]])
        else:
            row.append(str(performer[key]))
    ws.append(row)

# Create a new sheet for Worst Performers
wb.create_sheet(title="Worst Performers")
ws = wb["Worst Performers"]

# Dynamically write headers based on keys of the first entry in worst_performers
if worst_performers:
    headers = []
    for key in worst_performers[0].keys():
        if isinstance(worst_performers[0][key], list):
            # Expand array values into separate columns
            headers.extend([f"{key}_{i}" for i in range(len(worst_performers[0][key]))])
        else:
            headers.append(key)
    ws.append(headers)

# Write data rows for Worst Performers
for performer in worst_performers:
    row = []
    for key in performer.keys():
        if isinstance(performer[key], list):
            # Expand array values into separate columns
            row.extend([str(item) for item in performer[key]])
        else:
            row.append(str(performer[key]))
    ws.append(row)

# Create a new sheet for Hits Sem Scores
wb.create_sheet(title="Hits_1 Sem Scores")
ws = wb["Hits_1 Sem Scores"]
# Write headers for Hits Sem Scores
ws.append(["Hit_1/Not Hit_1", "Mean semscore", "Variance"])
# Write data for Hits Sem Scores
for key, sem_scores in hits_sem_scores.items():
    mean = calculate_mean(sem_scores)
    variance = calculate_variance(sem_scores)
    ws.append([key, mean, variance])

# Save the workbook
# wb.save("sem_scores_analysis.xlsx")

# Flatten all sem_scores and round to two decimals
all_sem_scores = flatten([flatten(scores) for scores in sem_scores_obj.values()])
rounded_scores = [round(score, 2) for score in all_sem_scores]

# Count frequencies of each rounded score
score_counts = Counter(rounded_scores)

# Sort scores for plotting
sorted_scores = sorted(score_counts.items())
x_values, y_values = zip(*sorted_scores)

# Plot the distribution graph
plt.figure(figsize=(10, 6))
plt.bar(x_values, y_values, width=0.01, color='blue', alpha=0.7)
plt.xlabel("Sem Score (rounded to 2 decimals)")
plt.ylabel("Frequency")
plt.title("Distribution of Sem Scores")
plt.grid(axis='y', linestyle='--', alpha=0.7)
plt.tight_layout()

# Save the plot as an image or display it
plt.savefig("sem_score_distribution.png")
plt.show()
